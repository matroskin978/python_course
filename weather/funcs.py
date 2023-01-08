import datetime
import config
import requests
from os import path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def get_date_time(ts, timezone, dt_format="%H:%M:%S"):
    tz = datetime.timezone(datetime.timedelta(seconds=timezone))
    return datetime.datetime.fromtimestamp(ts, tz=tz).strftime(dt_format)


def get_weather(city_name):
    params = {
        "appid": config.API_KEY,
        "units": config.UNITS,
        "lang": config.LANG,
        "q": city_name
    }
    try:
        r = requests.get(config.API_URL, params=params)
        return r.json()
    except:
        return {"cod": 0, "message": "Не удалось получить данные"}


def print_weather(data):
    if data['cod'] != 200:
        print(data['message'])
        return {}
    else:
        # print(data)
        sunrise_time = get_date_time(data['sys']['sunrise'], data['timezone'])
        sunset_time = get_date_time(data['sys']['sunset'], data['timezone'])

        print(f"""
Местоположение: {data['name']}, {data['sys']['country']}
Температура: {data['main']['temp']} °C
Атм. давление: {data['main']['pressure']} гПа
Влажность: {data['main']['humidity']}%
Скорость ветра: {data['wind']['speed']} м/с
Погодные условия: {data['weather'][0]['description']}
Восход: {sunrise_time}
Закат: {sunset_time}
""")
        print('+' * 50)
        return data


def save_excel(data):
    if data['cod'] == 200:
        if path.exists(config.FILE_EXCEL):
            wb = load_workbook(filename=config.FILE_EXCEL)
            ws1 = wb.active
        else:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'Статистика запросов'
            ws1.append(['Дата запроса', 'Город', 'Температура °C'])
            ft = Font(color='FF0000', bold=True)
            a1 = ws1['A1']
            b1 = ws1['B1']
            c1 = ws1['C1']
            a1.font = ft
            b1.font = ft
            c1.font = ft

        ws1.append([datetime.datetime.now(), f"{data['name']}, {data['sys']['country']}", data['main']['temp']])
        wb.save(filename=config.FILE_EXCEL)
