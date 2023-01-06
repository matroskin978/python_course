# https://openweathermap.org/current
# https://pypi.org/project/requests/
import requests
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from os import path

API_KEY = '5e1ce895b1f7950c8267adecc8ce4989'
API_URL = 'https://api.openweathermap.org/data/2.5/weather'
UNITS = 'metric'
LANG = 'ru'
CITY = 'Лондон'
FILE_EXCEL = 'data.xlsx'


def get_date_time(ts, timezone, dt_format="%H:%M:%S"):
    tz = datetime.timezone(datetime.timedelta(seconds=timezone))
    return datetime.datetime.fromtimestamp(ts, tz=tz).strftime(dt_format)


def get_weather(city_name):
    params = {
        "appid": API_KEY,
        "units": UNITS,
        "lang": LANG,
        "q": city_name
    }
    try:
        r = requests.get(API_URL, params=params)
        return r.json()
    except:
        return {"cod": 0, "message": "Не удалось получить данные"}


def print_weather(data):
    if data['cod'] != 200:
        print(data['message'])
        return {}
    else:
        # print(data)
        sunrise_time = get_date_time(data['sys']['sunrise'], data['timezone'])
        sunset_time = get_date_time(data['sys']['sunset'], data['timezone'])

        print(f"""
Местоположение: {data['name']}, {data['sys']['country']}
Температура: {data['main']['temp']} °C
Атм. давление: {data['main']['pressure']} гПа
Влажность: {data['main']['humidity']}%
Скорость ветра: {data['wind']['speed']} м/с
Погодные условия: {data['weather'][0]['description']}
Восход: {sunrise_time}
Закат: {sunset_time}
""")
        print('+' * 50)
        return data


def save_excel(data):
    if data['cod'] == 200:
        if path.exists(FILE_EXCEL):
            wb = load_workbook(filename=FILE_EXCEL)
            ws1 = wb.active
        else:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'Статистика запросов'
            ws1.append(['Дата запроса', 'Город', 'Температура °C'])
            ft = Font(color='FF0000', bold=True)
            a1 = ws1['A1']
            b1 = ws1['B1']
            c1 = ws1['C1']
            a1.font = ft
            b1.font = ft
            c1.font = ft

        ws1.append([datetime.datetime.now(), f"{data['name']}, {data['sys']['country']}", data['main']['temp']])
        wb.save(filename=FILE_EXCEL)


print("*" * 70)
print("""* Привет! Я помогу узнать погоду в любом городе мира.
* Просто введи запрос в формате city[,country_code]
* Если нужно выйти из программы, тогда просто нажми Enter""")
print("*" * 70)

while True:
    q = input('Введи название города: ')
    if not q:
        print('До новых встреч!')
        break
    else:
        weather = get_weather(q)
        print_weather(weather)
        save_excel(weather)
